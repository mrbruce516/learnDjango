# 部门管理视图
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.views.decorators.csrf import csrf_exempt
from openpyxl import load_workbook
from web import models
from web.utils.pagination import Pagination
from web.utils.search import Search
from web.utils.form import DepartModelForm


# Create your views here.

def depart_list(request):
    """ 部门管理中心 """

    query_title = Search(request, field_rule="title__contains")

    # 去数据库中获取所有部门信息
    depart = models.Department.objects.filter(**query_title.data_dict)

    page_obj = Pagination(request, queryset=depart)

    context = {
        'search': query_title.search,
        'depart': page_obj.page_queryset,
        'placeholder': "查找部门",
        'page_number': page_obj.html()
    }

    return render(request, 'depart_list.html', context)


def depart_add(request):
    if request.method == 'GET':
        context = {
            'title': "新增部门",
            'form': DepartModelForm()
        }
        return render(request, 'add.html', context)

    form = DepartModelForm(data=request.POST)
    context = {
        'title': "新增部门",
        'form': form
    }
    if form.is_valid():
        form.save()
        # 重定向回主页面
        return redirect("/depart/list/")
    return render(request, 'add.html', context)


def depart_del(request):
    nid = request.GET.get("nid")
    models.Department.objects.filter(id=nid).delete()
    return redirect("/depart/list")


def depart_edit(request, nid):
    row_obj = models.Department.objects.filter(id=nid).first()
    if request.method == 'GET':
        context = {
            'title': "编辑部门",
            'form': DepartModelForm(instance=row_obj)
        }
        return render(request, "edit.html", context)

    # 获取数据并更新
    form = DepartModelForm(data=request.POST, instance=row_obj)
    if form.is_valid():
        form.save()
        return redirect('/depart/list/')
    else:
        context = {
            'title': "编辑部门",
            'form': form
        }
        return render(request, "edit.html", context)


@csrf_exempt
def depart_batch_add(request):
    # 获取excel文件
    file_obj = request.FILES.get('file')
    # 把用户的excel文件读到内存中
    wb = load_workbook(file_obj)
    sheet = wb.worksheets[0]
    # 循环获取每一行数据
    for row in sheet.iter_rows(min_row=2):
        data = row[0].value
        # 添加到数据库
        exists = models.Department.objects.filter(title=data).exists()
        if not exists:
            models.Department.objects.create(title=data)
        return HttpResponse("数据重复，添加失败")
    return HttpResponse("上传成功")
